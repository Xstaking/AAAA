import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
from tkinter import messagebox, simpledialog
import os

EXCEL_FILE = "C:\\Users\\Angelo\\Downloads\\final\\inventory.xlsx"
bg_file = "C:\\Users\\Angelo\\Downloads\\final\\bg.png" #CHANGE THIS INTO YOUR DIRECTORY OF THE FILE AND MAKE SURE TO USE '\\'

#---- VARIABLES FOR WORKSHEET TITLES ------
# 'login_ws' for login history
# 'userSheets' for RegisteredUsers
# 'inventory_ws' for Inventory


def log_login_time(username):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = Workbook()
        login_ws = wb.active
        login_ws.title = "LoginHistory"
        login_ws.append(["Username", "Timestamp"])
        wb.save(EXCEL_FILE)

    if "LoginHistory" not in wb.sheetnames:
        login_ws = wb.create_sheet("LoginHistory")
        login_ws.append(["Username", "Timestamp"])
    else:
        login_ws = wb["LoginHistory"]

    login_ws.append([username, now])
    wb.save(EXCEL_FILE)
    wb.close()

def init_credentials_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        userSheets = wb.active
        userSheets.title = "RegisteredUsers"
        userSheets.append(["Username", "Password"])
        userSheets.append(["admin", "admin123"])
        wb.save(EXCEL_FILE)

def validate_login(username, password):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            wb.close()
            return True
    wb.close()
    return False

def user_exists(username):
    wb = load_workbook(EXCEL_FILE)
    userSheets = wb.active
    for row in userSheets.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            wb.close()
            return True
    wb.close()
    return False

# def show_register_page():
#     if 'login_frame' in globals() and 'register_frame' in globals():
#         login_frame.grid_forget()
#         register_frame.grid()
#     else:
#         messagebox.showerror("Error", "Frames are not initialized.")

    if not user or not pwd:
        messagebox.showwarning("Input Error", "Please fill in both fields.")
        return

    # Check if the user exists in the "RegisteredUsers" worksheet
    wb = load_workbook(EXCEL_FILE)
    if "RegisteredUsers" not in wb.sheetnames:
        userSheets = wb.create_sheet("RegisteredUsers")
        userSheets.append(["Username", "Password"])  # Add headers
        wb.save(EXCEL_FILE)
    else:
        userSheets = wb["RegisteredUsers"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user:
            wb.close()
            messagebox.showerror("Error", "Username already exists.")
            return

    # Register the user in the "RegisteredUsers" worksheet
    ws.append([user, pwd])
    wb.save(EXCEL_FILE)
    wb.close()

    messagebox.showinfo("Success", "User registered successfully!")
    show_login_page()

def register_user(username, password):
    wb = load_workbook(EXCEL_FILE)
    
    # Check if "RegisteredUsers" sheet exists, otherwise create it
    if "RegisteredUsers" not in wb.sheetnames:
        userSheets = wb.create_sheet("RegisteredUsers")
        userSheets.append(["Username", "Password"])
        
    else:
        userSheets = wb["RegisteredUsers"]
    
    userSheets.append([username, password])
    wb.save(EXCEL_FILE)
    wb.close()

# Initialize global variables
current_user = None

def login():
    global current_user
    user = login_username.get().strip()
    pwd = login_password.get().strip()

    if not user or not pwd:
        messagebox.showwarning("Input Error", "Please fill in both fields.")
        return

    # Validate login against the "RegisteredUsers" sheet
    try:
        wb = load_workbook(EXCEL_FILE)
        if "RegisteredUsers" in wb.sheetnames:
            userSheets = wb["RegisteredUsers"]
            for row in userSheets.iter_rows(min_row=2, values_only=True):
                if row[0] == user and row[1] == pwd:
                    current_user = user
                    log_login_time(user)
                    messagebox.showinfo("Login Success", f"Welcome, {user}!")
                    inventoryWindow()
                    wb.close()
                    return
        wb.close()
    except FileNotFoundError:
        messagebox.showerror("Error", "Credentials file not found.")

    messagebox.showerror("Login Failed", "Invalid credentials.")

def show_login_page():
    if 'register_frame' in globals() and 'login_frame' in globals():
        login_frame.grid()
    else:
        messagebox.showerror("Error", "Frames are not initialized.")

def getAdminPassword(EXCEL_FILE, admin_username="admin"):
    wb = load_workbook(EXCEL_FILE)
    userSheets = wb.active

    for row in userSheets.iter_rows(values_only=True):
        if row[0] == admin_username:  #column A
            return row[1]  #column B
    return None

def inventoryWindow():
    root.withdraw()
    adminPassword = getAdminPassword(EXCEL_FILE)
    def logout():
        answer = messagebox.askyesno("Confirm Logout", "Are you sure you want to log out?")
        if answer:
            messagebox.showinfo("Logout", "You have been logged out.")
            Management.destroy()
            root.deiconify()
        else:
            # Redirect to a valid tab (e.g., Dashboard)
            notebook.select(0)
    def on_tab_changed(event):
        selected_tab = event.widget.index("current")
        if selected_tab == 4:
            if not access_granted[0]:
                pwd = simpledialog.askstring("Password Required", "Enter password:", show="*")
                if pwd == adminPassword:
                    access_granted[0] = True
                    # Allow access to tab 4
                else:
                    messagebox.showerror("Access Denied", "Incorrect password.")
                    notebook.select(0)  # Revert to first tab
                
                def register():
                    user = reg_username.get().strip()
                    pwd = reg_password.get().strip()

                    if not user or not pwd:
                        messagebox.showwarning("Input Error", "Please fill in both fields.")
                        return

                    # Check if the user exists in the "RegisteredUsers" worksheet
                    wb = load_workbook(EXCEL_FILE)
                    if "RegisteredUsers" not in wb.sheetnames:
                        userSheets = wb.create_sheet("RegisteredUsers")
                        userSheets.append(["Username", "Password"])  # Add headers
                        wb.save(EXCEL_FILE)
                    else:
                        userSheets = wb["RegisteredUsers"]

                    for row in userSheets.iter_rows(min_row=2, values_only=True):
                        if row[0] == user:
                            wb.close()
                            messagebox.showerror("Error", "Username already exists.")
                            return True

                    # Register the user in the "RegisteredUsers" worksheet
                    userSheets.append([user, pwd])
                    wb.save(EXCEL_FILE)
                    wb.close()

                    messagebox.showinfo("Success", "User registered successfully!")

            
                
                # TAb 5

                def load_userList(EXCEL_FILE, user_tree):
                    for col in user_tree["columns"]:
                        user_tree.heading(col, text="")
                        user_tree.delete(*user_tree.get_children())

                    wb = load_workbook(EXCEL_FILE)
                    userSheets = wb["RegisteredUsers"]

                    headers = []
                    for cell in userSheets[1]:
                        headers.append(cell.value)

                    user_tree["columns"] = headers
                    user_tree["show"] = "headings"

                    for col in headers:
                        user_tree.heading(col, text=col)
                        user_tree.column(col, width=100)  # You can adjust width as needed

                    # Insert the rest of the rows
                    for row in userSheets.iter_rows(min_row=2, values_only=True):
                        user_tree.insert("", "end", values=row)

                register_frame = tk.Frame(tab5, bg="gray10")
                register_frame.grid(row=0, column=0, sticky="nsew")

                tk.Label(tab5, text="Register", font=("Arial", 32, "bold"), bg="gray10", fg="white").grid(row=0, column=0, padx=15, pady=15, sticky="w")
                tk.Label(tab5, text="Username", bg="gray10", font=("Arial", 12), fg="white").grid(row=1, column=0, padx=15, pady=10, sticky="w")
                reg_username = tk.Entry(tab5, font=("Arial", 12))
                reg_username.grid(row=1, column=1, padx=15, pady=10, sticky="we")
                tk.Label(tab5, text="Password", bg="gray10", font=("Arial", 12), fg="white").grid(row=2, column=0, padx=15, pady=10, sticky="w")
                reg_password = tk.Entry(tab5, show='*', font=("Arial", 12))
                reg_password.grid(row=2, column=1, padx=15, pady=10, sticky="we")

                def backToLogin(): #For "Back to Login Button"
                    Management.destroy()
                    root.deiconify()
                    
                # Transparent buttons
                tk.Button(tab5, text="SUBMIT", command=register, font=("Arial", 12), bg="gray10", fg="white", relief="groove").grid(row=1, column=3, columnspan=2, pady=15)
                tk.Button(tab5, text="BACK TO LOGIN", command=backToLogin, font=("Arial", 12), bg="gray10", fg="white", relief="groove").grid(row=2, column=3, columnspan=2)
    
                def delete_item_user(user_tree):
                    selected = user_tree.selection()
                    if not selected:
                        messagebox.showwarning("Warning", "No item selected.")
                        return

                    # Get the selected item values
                    selected_values = user_tree.item(selected[0], "values")
                    if not selected_values:
                        messagebox.showerror("Error", "Could not retrieve selected values.")
                        return
                    
                    confirm = messagebox.askyesno(
                        "Confirm Delete",
                        f"Are you sure you want to delete this entry?\n\n{selected_values}"
                    )
                    if not confirm:
                        return

                    # Load workbook and sheet
                    wb = load_workbook(EXCEL_FILE)
                    userSheets = wb["RegisteredUsers"]

                    found = False
                    for row_idx, row in enumerate(userSheets.iter_rows(min_row=2), start=2):
                        excel_values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
                        selected_values_str = [str(val).strip() if val is not None else "" for val in selected_values]

                        if excel_values == selected_values_str:
                            userSheets.delete_rows(row_idx, 1)
                            found = True
                            break

                    if found:
                        wb.save(EXCEL_FILE)
                        user_tree.delete(selected[0])
                        messagebox.showinfo("Success", "Entry deleted from Excel and Treeview.")
                    else:
                        messagebox.showwarning("Not Found", "Matching entry not found in Excel.")

                def update_item_user(user_tree):
                    selected = user_tree.focus()
                    if not selected:
                        print("No item selected.")
                        return
                    
                    values = user_tree.item(selected, 'values')
                    columns = user_tree["columns"]

                    username_index = columns.index('Username')
                    original_username = values[username_index]

                    edit_win = tk.Toplevel()
                    edit_win.title("Edit Credentials")
                    entry_vars = {}


                    for i, col in enumerate(columns):
                        tk.Label(edit_win, text=col.capitalize()).grid(row=i, column=0, padx=5, pady=2, sticky="w")
                        var = tk.StringVar(value=values[i])
                        tk.Entry(edit_win, textvariable=var).grid(row=i, column=1, padx=5, pady=2)
                        entry_vars[col] = var
                    
                    def save_credentials():
                        new_values = [entry_vars[col].get() for col in columns]
                        user_tree.item(selected, values=new_values)

                        wb = load_workbook(EXCEL_FILE)
                        userSheets = wb.active

                        # Find and update the row using original username
                        for row in userSheets.iter_rows(min_row=2):
                            if str(row[username_index].value) == original_username:
                                for j, col in enumerate(columns):
                                    row[j].value = new_values[j]
                                break

                        wb.save(EXCEL_FILE)
                        edit_win.destroy()
                    tk.Button(edit_win, text="Save", command=save_credentials).grid(row=len(columns), column=0, columnspan=2, pady=10)

                user_tree = ttk.Treeview(tab5, columns=("Username", "Password"), show="headings")
                style = ttk.Style()
                style.theme_use("clam")
                style.configure("Treeview.Heading", foreground="white", background="DarkSlateGrey")
                user_tree.grid(row=5, column=0, columnspan=5, sticky="nsew")
                load_userList(EXCEL_FILE, user_tree)

                tab5.grid_columnconfigure(0, weight=1)  
                tab5.grid_columnconfigure(1, weight=2)
                tab5.grid_columnconfigure(2, weight=1)
                tab5.grid_columnconfigure(3, weight=1)
                tab5.grid_columnconfigure(4, weight=1)
                tab5.grid_rowconfigure(7, weight=1)

                for col in ("Username", "Password"):
                    user_tree.heading(col, text=col)
                    user_tree.grid(row=5, column=0, columnspan=5)

                    try:
                        wb = load_workbook("inventory.xlsx")
                    except FileNotFoundError:
                        wb = Workbook()
                        userSheets = wb.active
                        userSheets.append("Username", "Password")  
                        wb.save("inventory.xlsx")

                        ws = wb.active
                        selected_item_index = None 
    
                tk.Button(tab5, text="REFRESH", command=lambda: load_userList(EXCEL_FILE, user_tree), bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=4, column=0, pady=5, sticky="we")
                tk.Button(tab5, text="DELETE", command=lambda: delete_item_user(user_tree), bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=4, column=1, pady=5, sticky="we")
                tk.Button(tab5, text="UPDATE", command=lambda: update_item_user(user_tree), bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=4, column=2, pady=5, sticky="we")
        elif selected_tab == 5:
            logout()

    access_granted = [False]
    
    Management = tk.Toplevel(root)
    Management.title("Inventory Management")
    Management.configure(bg="gray10")

    style = ttk.Style()
    style.configure('top.TNotebook', tabposition='n')  # Tabs on the left
    notebook = ttk.Notebook(Management, style='top.TNotebook')

    # Create a frame for each tab
    tab1 = tk.Frame(notebook)
    tab2 = tk.Frame(notebook, bg="gray10")
    tab3 = tk.Frame(notebook, bg="gray10")
    tab4 = tk.Frame(notebook, bg="gray10")
    tab5 = tk.Frame(notebook, bg="gray10")
    tab6 = tk.Frame(notebook, bg="gray10", width=1, height=1)

    #Load and display background image
    bg_img = tk.PhotoImage(file=bg_file)
    #bg_img = bg_img.subsample(2,2)
    tab1.bg_img = bg_img
    bg_label = tk.Label(tab1, image=bg_img, bg="black")  # Set background color to black
    bg_label.place(x=0, y=0, relwidth=1, relheight=1)

    # Add tabs to notebook (each only once)
    notebook.add(tab1, text="   HOME   ")
    notebook.add(tab2, text="   INPUT   ")
    notebook.add(tab3, text="   HISTORY   ")
    notebook.add(tab4, text="   INBOX   ")
    notebook.add(tab5, text="   USER MANAGEMENT   ")
    notebook.add(tab6, text="   LOGOUT   ")

    # Style config
    style.theme_use("clam")
    style.configure("TNotebook", background="DarkSlateGrey", borderwidth=0)
    style.configure("TNotebook.Tab", background="DarkSlateGrey", foreground="white", padding=[10, 5], font=("Arial", 10, "bold"))
    style.map("TNotebook.Tab", background=[("selected", "gray10")], foreground=[("selected", "white")])

    # Layout
    notebook.grid(sticky="nsew")
    style.configure('top.TNotebook', tabposition='n')  # Set tabs to the left side
    notebook.bind("<<NotebookTabChanged>>", on_tab_changed)

    # Make the notebook stretchable
    Management.grid_columnconfigure(0, weight=1)
    Management.grid_rowconfigure(0, weight=1)


    wb = load_workbook(EXCEL_FILE)
    if "Inventory" not in wb.sheetnames:
        inventory_ws = wb.create_sheet("Inventory")
        inventory_ws.append(["Item", "Quantity", "Price"])  # Add headers
        wb.save(EXCEL_FILE)
    else:
        inventory_ws = wb["Inventory"]
        history_ws = wb["History"]

    inventory_ws = wb["Inventory"]
    history_ws = wb["History"]
    selected_item_index = None 
    def add_item():
        item, quantity, price = entry_item.get(), entry_quantity.get(), entry_price.get()
        if item and quantity and price:
            inventory_ws.append([item, quantity, price])  # Append the new row at the end
            wb.save("inventory.xlsx")
            inventoryTree.insert("", "end", values=(item, quantity, price))  # Insert at the end of the Treeview
            log_action("Add", item, quantity, price)  # Log action to history
            log_action_to_inbox("Add", item, quantity, price)  # Log action to inbox
            clear_entries()
            update_total()  # Update total after adding an item

    def view_data():
        wb = load_workbook(EXCEL_FILE)
        inventory_ws = wb["Inventory"]
        history_ws = wb["History"]

        # Build a dictionary for the most recent timestamp per item
        timestamp_dict = {}
        for row in history_ws.iter_rows(min_row=2, values_only=True):
            item = row[1]
            timestamp = row[4]
            if item and timestamp:
                if isinstance(timestamp, str):
                    try:
                        timestamp = datetime.fromisoformat(timestamp)
                    except ValueError:
                        continue  # Skip invalid timestamp
                if item not in timestamp_dict or timestamp > timestamp_dict[item]:
                    timestamp_dict[item] = timestamp

        # Clear the Treeview
        inventoryTree.delete(*inventoryTree.get_children())

        # Insert combined data: item, quantity, price, timestamp
        for row_index, row in enumerate(inventory_ws.iter_rows(min_row=2, values_only=True), start=2):
            item, quantity, price = row
            timestamp = timestamp_dict.get(item)
            formatted_time = timestamp.strftime("%Y-%m-%d %H:%M:%S") if timestamp else "N/A"
            inventoryTree.insert("", "end", values=(item, quantity, price, formatted_time), iid=str(row_index))


    def delete_item():
        global selected_item_index
        selected = inventoryTree.selection()
        if selected:
            row_index = int(selected[0])  # Row index
            values = inventoryTree.item(selected[0], "values")
            inventory_ws.delete_rows(row_index)  
            wb.save("inventory.xlsx")
            inventoryTree.delete(selected[0])  
            log_action("Delete", values[0], values[1], values[2])  # Log action to history
            log_action_to_inbox("Delete", values[0], values[1], values[2])  # Log action to inbox
            selected_item_index = None  # Reset after deletion
            update_total()  # Update total after deleting an item

    def edit_item():
        global selected_item_index
        selected = inventoryTree.selection()
        if selected:
            selected_item_index = int(selected[0])  # Store row index
            values = inventoryTree.item(selected[0], "values")
            entry_item.delete(0, tk.END)
            entry_item.insert(0, values[0])
            entry_quantity.delete(0, tk.END)
            entry_quantity.insert(0, values[1])
            entry_price.delete(0, tk.END)
            entry_price.insert(0, values[2])

    def update_item():
        global selected_item_index
        if selected_item_index:
            new_values = (entry_item.get(), entry_quantity.get(), entry_price.get())
            if all(new_values):
                inventoryTree.item(str(selected_item_index), values=new_values) 
                inventory_ws[selected_item_index][0].value = new_values[0]  
                inventory_ws[selected_item_index][1].value = new_values[1]
                inventory_ws[selected_item_index][2].value = new_values[2]
                wb.save("inventory.xlsx")
                log_action("Update", new_values[0], new_values[1], new_values[2])  # Log action to history
                log_action_to_inbox("Update", new_values[0], new_values[1], new_values[2])  # Log action to inbox
                clear_entries()
                selected_item_index = None 
                update_total()  # Update total after updating an item

    def clear_entries():
        entry_item.delete(0, tk.END)
        entry_quantity.delete(0, tk.END)
        entry_price.delete(0, tk.END)

 
    #Tab 2 - Input

    def load_inventoryList(EXCEL_FILE, user_tree):
        for col in user_tree["columns"]:
            user_tree.heading(col, text="")
            user_tree.delete(*user_tree.get_children())

            wb = load_workbook(EXCEL_FILE)
            userSheets = wb["RegisteredUsers"]

            headers = []
            for cell in userSheets[1]:
                headers.append(cell.value)

                user_tree["columns"] = headers
                user_tree["show"] = "headings"

                for col in headers:
                    user_tree.heading(col, text=col)
                    user_tree.column(col, width=100)  # You can adjust width as needed

                # Insert the rest of the rows
                    for row in userSheets.iter_rows(min_row=2, values_only=True):
                        user_tree.insert("", "end", values=row)

    tk.Label(tab2, text="Input", font=("Arial", 32, "bold"), bg="gray10", fg="white").grid(row=0, column=0, padx=15, pady=15, sticky="w")
    tk.Label(tab2, text="Item", bg="gray10", font="Calibri 11", fg="white").grid(row=1, column=0, sticky="w")
    tk.Label(tab2, text="Quantity", bg="gray10", font="Calibri 11", fg="white").grid(row=2, column=0, sticky="w")
    tk.Label(tab2, text="Price", bg="gray10", font="Calibri 11", fg="white").grid(row=3, column=0, sticky="w")

    entry_item, entry_quantity, entry_price = tk.Entry(tab2, font="Times 11", justify="left"), tk.Entry(tab2, font="Times 11", justify="left"), tk.Entry(tab2, font="Times 11", justify="left")

    entry_item.grid(row=1, column=1, sticky="we")
    entry_quantity.grid(row=2, column=1,  sticky="we")
    entry_price.grid(row=3, column=1, sticky="we")

    tk.Button(tab2, text="ADD", command=add_item, bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=5, column=0, pady=5, sticky="we")
    tk.Button(tab2, text="REFRESH", command=view_data, bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=5, column=1, pady=5, sticky="we")
    tk.Button(tab2, text="EDIT", command=edit_item, bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=5, column=2, pady=5, sticky="we")
    tk.Button(tab2, text="DELETE", command=delete_item, bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=5, column=3, pady=5, sticky="we")
    tk.Button(tab2, text="UPDATE", command=update_item, bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=5, column=4, pady=5, sticky="we")

    inventoryTree = ttk.Treeview(tab2, columns=("Item", "Quantity", "Price", "Date Added"), show="headings")
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview.Heading", foreground="white", background="DarkSlateGrey")
    inventoryTree.grid(row=8, column=0, columnspan=5, sticky="nsew")



    tab2.grid_columnconfigure(0, weight=1)  
    tab2.grid_columnconfigure(1, weight=2)
    tab2.grid_columnconfigure(2, weight=1)
    tab2.grid_columnconfigure(3, weight=1)
    tab2.grid_columnconfigure(4, weight=1)
    tab2.grid_rowconfigure(8, weight=1)

    for col in ("Item", "Quantity", "Price", "Date Added"):
        inventoryTree.heading(col, text=col)
    inventoryTree.grid(row=8, column=0, columnspan=5)

    def search(event):
        timestamp_lookup = {}
        for row in history_ws.iter_rows(min_row=2, values_only=True):
            try:
                item = row[1] 
                timestamp = row[4]
                if item:
                    timestamp_lookup[str(item).lower()] = timestamp
            except IndexError:
                continue
        
        query = entry.get().lower()
        inventoryTree.delete(*inventoryTree.get_children()) 
        total = 0
        searched = bool(query)
        for row in inventory_ws.iter_rows(min_row=2, values_only=True):
            try:
                item, quantity, price = row[:3]
                if not item:
                    continue

                item_str = str(item).lower()
                timestamp = timestamp_lookup.get(item_str, "N/A")

                if query in item_str or not searched:
                    inventoryTree.insert("", "end", values=(item, quantity, price, timestamp))
                    if searched and query in item_str:
                        total += float(quantity) * float(price)
            except (ValueError, TypeError, IndexError):
                continue

        total_var.set(str(round(total, 2)) if searched else "0")

    label = tk.Label(tab2, text="Search", bg="gray10", font="Times 11", fg="white")
    label.grid(row=7, column=0, pady=10, sticky="w")
    entry = tk.Entry(tab2, font="Times 11", justify="left")
    entry.grid(row=7, column=1, pady=5, sticky="we")
    entry.bind("<KeyRelease>", search)
    total_var = tk.StringVar()
    tk.Label(tab2, text="Total of the Selected item: ", bg="gray10", font="Calibri 11", fg="white").grid(row=7, column=2, sticky="w", padx =5)
    total_item = tk.Entry(tab2, font = "Times 11", justify = "left", textvariable=total_var, state = "readonly").grid(row = 7, column = 3, sticky = "we")

    tab2.grid_columnconfigure(1, weight=1)  
    view_data()

    


    def update_total():
        def deduct_quantity():
            selected = inventoryTree.selection()
            if selected:
                values = inventoryTree.item(selected[0], "values")
                item_name = values[0]
                current_quantity = values[1]

                try:
                    current_quantity = int(current_quantity)
                    deduction = int(spinbox_deduction.get())
                    if deduction > current_quantity:
                        messagebox.showerror("Error", "Deduction exceeds current quantity.")
                        return

                    new_quantity = current_quantity - deduction
                    inventory_ws[int(selected[0])][1].value = new_quantity  # Update quantity in Excel
                    wb.save("inventory.xlsx")
                    inventoryTree.item(selected[0], values=(item_name, new_quantity, values[2]))  # Update Treeview
                    update_total()  # Update total after deduction
                    messagebox.showinfo("Success", f"Deducted {deduction} from {item_name}.")
                    
                    # Log action to history and inbox
                    log_action("Deduct", item_name, deduction, values[2])
                    log_action_to_inbox("Deduct", item_name, deduction, values[2])
                except ValueError:
                    messagebox.showerror("Error", "Invalid quantity or deduction value.")
            else:
                messagebox.showwarning("Warning", "No item selected.")

        tk.Label(tab2, text="Deduct Quantity", bg="gray10", font="Calibri 11", fg="white").grid(row=6, column=0, sticky="w")
        spinbox_deduction = tk.Spinbox(tab2, from_=0, to=1000, increment=1, font="Times 11", justify="left")
        spinbox_deduction.grid(row=6, column=1, sticky="we")
        tk.Button(tab2, text="DEDUCT", command=deduct_quantity, bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=6, column=2, pady=5, sticky="we")
        total = 0
        for row_index, row in enumerate(inventory_ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[1] is not None and row[2] is not None:  
                try:
                    quantity = float(row[1])
                    price = float(row[2])
                    total += quantity * price
                except ValueError:
                    continue

        total_row = ["Total", "", f"{total:.2f}"]

        for row_index, row in enumerate(inventory_ws.iter_rows(min_row=2, max_row=inventory_ws.max_row, values_only=True), start=2):
            if row[0] == "Total":
                inventory_ws.delete_rows(row_index)
                break

        inventory_ws.append(total_row)
        wb.save("inventory.xlsx")

    
        inventoryTree.delete(*inventoryTree.get_children())
        for row_index, row in enumerate(inventory_ws.iter_rows(min_row=2, values_only=True), start=2):
            inventoryTree.insert("", "end", values=row, iid=str(row_index))

    
        view_data()

    tk.Label(tab3, text="History", font=("Arial", 32, "bold"), bg="gray10", fg="white").grid(row=0, column=0, padx=15, pady=15, sticky="w")

    history_tree = ttk.Treeview(tab3, columns=("Action", "Item", "Quantity", "Price", "Timestamp"), show="headings")
    history_tree.grid(row=1, column=0, columnspan=5, sticky="nsew")

    for col in ("Action", "Item", "Quantity", "Price", "Timestamp"):
        history_tree.heading(col, text=col)

    tab3.grid_columnconfigure(0, weight=1)
    tab3.grid_rowconfigure(1, weight=1)

    # Create a new worksheet for history if it doesn't exist
    if "History" not in wb.sheetnames:
        history_ws = wb.create_sheet("History")
        history_ws.append(["Action", "Item", "Quantity", "Price", "Timestamp"])
    else:
        history_ws = wb["History"]

    def log_action(action, item, quantity, price):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        history_ws.append([action, item, quantity, price, timestamp])
        wb.save("inventory.xlsx")
        history_tree.insert("", "end", values=(action, item, quantity, price, timestamp))
        messagebox.showinfo("Action Logged", f"{action} action performed on item '{item}'.")

    def add_item_with_logging():
        item, quantity, price = entry_item.get(), entry_quantity.get(), entry_price.get()
        add_item()  # Call the original add_item function
        #if item and quantity and price:
            #log_action("Add", item, quantity, price)

    def delete_item_with_logging():
        global selected_item_index
        selected = inventoryTree.selection()
        if selected:
            values = inventoryTree.item(selected[0], "values")
            delete_item()  # Call the original delete_item function
            #log_action("Delete", values[0], values[1], values[2])

    def update_item_with_logging():
        global selected_item_index
        if selected_item_index:
            new_values = (entry_item.get(), entry_quantity.get(), entry_price.get())
            update_item()  # Call the original update_item function
            #if all(new_values):
                #log_action("Update", new_values[0], new_values[1], new_values[2])
    
    tk.Button(tab2, text="ADD", command=add_item_with_logging, bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=5, column=0, pady=5, sticky="we")
    tk.Button(tab2, text="DELETE", command=delete_item_with_logging, bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=5, column=3, pady=5, sticky="we")
    tk.Button(tab2, text="UPDATE", command=update_item_with_logging, bg="DarkSlateGrey", fg="white", relief="groove", font="Arial 10").grid(row=5, column=4, pady=5, sticky="we")

    for row in history_ws.iter_rows(min_row=2, values_only=True):
        history_tree.insert("", "end", values=row)
    
    def search_history(event):
        query = history_entry.get().lower()
        history_tree.delete(*history_tree.get_children())  

        for row_index, row in enumerate(history_ws.iter_rows(min_row=2, values_only=True), start=2):
            if query in " ".join(map(str, row)).lower(): 
                history_tree.insert("", "end", values=row, iid=str(row_index))

    history_label = tk.Label(tab3, text="Search History", bg="gray10", font="Times 11", fg="white")
    history_label.grid(row=2, column=0, pady=10, sticky="w")
    history_entry = tk.Entry(tab3, font="Times 11", justify="left")
    history_entry.grid(row=2, column=1, columnspan=4, pady=5, sticky="we")
    history_entry.bind("<KeyRelease>", search_history)

    tab3.grid_columnconfigure(1, weight=1)

    view_data()
    update_total()

    # Tab 4 - Inbox
    tk.Label(tab4, text="Inbox", font=("Arial", 32, "bold"), bg="gray10", fg="white").grid(row=0, column=0, padx=15, pady=15, sticky="w")

    inbox_tree = ttk.Treeview(tab4, columns=("Sender", "Message", "Timestamp"), show="headings")
    inbox_tree.grid(row=1, column=0, columnspan=5, sticky="nsew")

    for col in ("Sender", "Message", "Timestamp"):
        inbox_tree.heading(col, text=col)

    # Create a new worksheet for Inbox if it doesn't exist
    if "Inbox" not in wb.sheetnames:
        inbox_ws = wb.create_sheet("Inbox")
        inbox_ws.append(["Sender", "Message", "Timestamp"])
        wb.save(EXCEL_FILE)
    else:
        inbox_ws = wb["Inbox"]

    # Function to display messages in the inbox treeview
    def display_inbox_message(sender, message, timestamp):
        inbox_tree.insert("", "end", values=(sender, message, timestamp))

    # Load initial inbox data
    def load_inbox_data():
        inbox_tree.delete(*inbox_tree.get_children())
        for row in inbox_ws.iter_rows(min_row=2, values_only=True):
            display_inbox_message(row[0], row[1], row[2])

    load_inbox_data()

    tab4.grid_columnconfigure(0, weight=1)
    tab4.grid_rowconfigure(1, weight=1)

    # Function to log messages to the Inbox worksheet
    def log_inbox_message(sender, message):
        global current_user
        sender = current_user if current_user else "System"
        timestamp = datetime.now().strftime("%d %b %Y, %I:%M %p")  # User-friendly timestamp format
        inbox_ws.append([sender, message, timestamp])
        wb.save("inventory.xlsx")
        display_inbox_message(sender, message, timestamp)

    # Automatically log messages to Inbox based on history actions
    def log_action_to_inbox(action, item, quantity, price):
        message = f"{action} performed on item '{item}' with quantity '{quantity}' and price '{price}'."
        log_inbox_message("System", message)

    # Update the history logging functions to also log to the Inbox
    def add_item_with_logging():
        item, quantity, price = entry_item.get(), entry_quantity.get(), entry_price.get()
        add_item()  # Call the original add_item function
        #if item and quantity and price:
            #log_action("Add", item, quantity, price)
            #log_action_to_inbox("Add", item, quantity, price)

    def delete_item_with_logging():
        global selected_item_index
        selected = inventoryTree.selection()
        if selected:
            values = inventoryTree.item(selected[0], "values")
            delete_item()  # Call the original delete_item function
            #log_action("Delete", values[0], values[1], values[2])
            #log_action_to_inbox("Delete", values[0], values[1], values[2])

    def update_item_with_logging():
        global selected_item_index
        if selected_item_index:
            new_values = (entry_item.get(), entry_quantity.get(), entry_price.get())
            update_item()  # Call the original update_item function
            #if all(new_values):
                #log_action("Update", new_values[0], new_values[1], new_values[2])
                #log_action_to_inbox("Update", new_values[0], new_values[1], new_values[2])

    # Automatically update the Inbox when a new entry is added to the History worksheet
    def sync_inbox_with_history():
        history_rows = list(history_ws.iter_rows(min_row=2, values_only=True))
        inbox_rows = list(inbox_ws.iter_rows(min_row=2, values_only=True))

        # Add missing history rows to the Inbox
        for row in history_rows[len(inbox_rows):]:
            action, item, quantity, price, timestamp = row
            log_action_to_inbox(action, item, quantity, price)

    # Call sync function whenever the history is updated
    sync_inbox_with_history()

    # Search functionality for the Inbox
    def search_inbox():
        query = inbox_entry.get().lower()
        inbox_tree.delete(*inbox_tree.get_children())

        for row in inbox_ws.iter_rows(min_row=2, values_only=True):
            if query in " ".join(map(str, row)).lower():
                inbox_tree.insert("", "end", values=row)

    inbox_label = tk.Label(tab4, text="Search Inbox", bg="gray10", font="Times 11", fg="white")
    inbox_label.grid(row=2, column=0, pady=10, sticky="w")
    inbox_entry = tk.Entry(tab4, font="Times 11", justify="left")
    inbox_entry.grid(row=2, column=1, columnspan=4, pady=5, sticky="we")
    inbox_entry.bind("<KeyRelease>", lambda _: search_inbox())


    
init_credentials_file()

root = tk.Tk()
root.title("Login & Admin Management")

root.configure(bg="gray10")
root.resizable(False, False)

# ---------------------- Login Frame ---------------------

login_frame = tk.Frame(root, bg="gray10")
login_frame.grid(row=0, column=0, sticky="nsew")
root.geometry("330x500")  # Adjust the window width to make it narrower

logo_img = tk.PhotoImage(file="C:\\Users\\Angelo\\Downloads\\final\\logo.png").subsample(6, 6)  # Adjust subsample values to make the logo very small
logo_label = tk.Label(login_frame, image=logo_img, bg="grey10")
logo_label.image = logo_img
logo_label.grid(row=0, column=0, columnspan=2, pady=(20, 10))

tk.Label(login_frame, text="POS Inventory", font=("Georgia", 28, "bold"), bg="grey10", fg="white").grid(row=1, column=0, columnspan=2, pady=(10, 5), padx=20)
tk.Label(login_frame, text="LOGIN PAGE", font=("Arial", 16), bg="grey10", fg="white").grid(row=2, column=0, columnspan=2, pady=(5, 20))

login_username = tk.Entry(login_frame, font=("Georgia", 12), fg="grey")
login_username.insert(0, "Username")
login_username.bind("<FocusIn>", lambda event: on_entry_click(event, login_username, "Username"))
login_username.bind("<FocusOut>", lambda event: on_focusout(event, login_username, "Username"))
login_username.grid(row=3, column=0, columnspan=2, padx=15, pady=10, sticky="we")

login_password = tk.Entry(login_frame, font=("Arial", 12), fg="grey", show="*")
login_password.insert(0, "Password")
login_password.bind("<FocusIn>", lambda event: on_entry_click(event, login_password, "Password"))
login_password.bind("<FocusOut>", lambda event: on_focusout(event, login_password, "Password"))
login_password.grid(row=4, column=0, columnspan=2, padx=15, pady=10, sticky="we")

tk.Button(
    login_frame, 
    text="LOGIN", 
    command=login, 
    font=("Arial", 12), 
    bg="DarkSlateGrey", 
    fg="white", 
    relief="groove", 
    highlightbackground="DarkSlateGrey", 
    highlightthickness=2, 
    bd=0, 
    padx=10, 
    pady=5, 
    borderwidth=2
).grid(row=5, column=0, columnspan=2, pady=(10, 20))

def on_entry_click(event, entry, placeholder):
    if entry.get() == placeholder:
        entry.delete(0, tk.END)
        entry.config(fg="black")

def on_focusout(event, entry, placeholder):
    if entry.get() == "":
        entry.insert(0, placeholder)
        entry.config(fg="grey")

root.mainloop()
